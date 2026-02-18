"""Export endpoints (Excel download)."""

import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..database import get_db
from .. import crud

router = APIRouter(prefix="/api/export", tags=["export"])


@router.get("/excel/{snapshot_id}")
def export_snapshot_excel(snapshot_id: int, db: Session = Depends(get_db)):
    snapshot = crud.get_snapshot(db, snapshot_id)
    if not snapshot:
        raise HTTPException(404, "Snapshot not found")

    import openpyxl
    wb = openpyxl.Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Snapshot ID", snapshot.id])
    ws.append(["Label", snapshot.label])
    ws.append(["Asset ID", snapshot.asset_id])
    ws.append(["Discount Rate", snapshot.discount_rate])
    ws.append(["Launch Year", snapshot.launch_year])
    ws.append(["Patent Expiry", snapshot.patent_expiry_year])
    ws.append(["Peak Sales ($M)", snapshot.peak_sales_usd_m])
    ws.append(["Uptake Curve", snapshot.uptake_curve])

    # Phase inputs sheet
    ws2 = wb.create_sheet("Phase Inputs")
    ws2.append(["Phase", "POS", "Duration (yr)", "Start Year"])
    for pi in snapshot.phase_inputs:
        ws2.append([pi.phase_name, pi.probability_of_success, pi.duration_years, pi.start_year])

    # R&D costs sheet
    ws3 = wb.create_sheet("R&D Costs")
    ws3.append(["Year", "Cost ($M)"])
    for rc in sorted(snapshot.rd_costs, key=lambda x: x.year):
        ws3.append([rc.year, rc.cost_usd_m])

    # Cashflows sheet
    ws4 = wb.create_sheet("Cashflows")
    ws4.append([
        "Year", "R&D Cost ($M)", "Commercial CF ($M)", "Net CF ($M)",
        "Cum POS", "Risk-Adj CF ($M)", "Discount Factor", "PV ($M)", "Cum NPV ($M)"
    ])
    for cf in sorted(snapshot.cashflows, key=lambda x: x.year):
        ws4.append([
            cf.year, cf.rd_cost_usd_m, cf.commercial_cf_usd_m, cf.net_cashflow_usd_m,
            cf.cumulative_pos, cf.risk_adjusted_cf_usd_m, cf.discount_factor,
            cf.pv_usd_m, cf.cumulative_npv_usd_m
        ])

    # Commercial rows sheet
    ws5 = wb.create_sheet("Commercial")
    ws5.append(["Year", "Gross Sales", "Net Sales", "COGS", "SG&A", "Op Profit", "Tax", "Net CF"])
    for cr in sorted(snapshot.commercial_rows, key=lambda x: x.year):
        ws5.append([
            cr.year, cr.gross_sales_usd_m, cr.net_sales_usd_m, cr.cogs_usd_m,
            cr.sga_usd_m, cr.operating_profit_usd_m, cr.tax_usd_m, cr.net_cashflow_usd_m
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"snapshot_{snapshot_id}_{snapshot.label.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
